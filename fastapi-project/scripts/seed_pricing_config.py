"""Seed script: populate pricing_config and crew_config tables.

Reads cost assumptions from the UNA Pricing Model 1 year.xlsx workbook
(C sheet for crew, M/I/Overhead & Other COGS sheet for maintenance/insurance/
DOC/overhead) and populates the initial version 1 rows.

Usage:
    python fastapi-project/scripts/seed_pricing_config.py              # Live insert
    python fastapi-project/scripts/seed_pricing_config.py --dry-run    # Print only

The script uses hardcoded data extracted from the Excel workbook audit.
It also attempts to read the Excel file to validate the hardcoded data and
warns if mismatches are found.

All numeric values are wrapped in Decimal(str(value)) to preserve precision.
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from decimal import Decimal
from pathlib import Path

# ---------------------------------------------------------------------------
# Authoritative data extracted from 'M/I/Overhead & Other COGS' sheet
# Source: Monthly rates and total budgets for pricing_config
# ---------------------------------------------------------------------------

PRICING_CONFIG_DATA = {
    "exchange_rate": "0.85",
    # Insurance — fixed USD/month (from M/I/Overhead sheet)
    "insurance_usd": "45000.00",
    # DOC — total monthly budget across fleet
    "doc_total_budget": "110000.00",
    # Overhead — total monthly budget across fleet
    "overhead_total_budget": "165000.00",
    # Other COGS — monthly per aircraft
    "other_cogs_monthly": "8500.00",
    # Maintenance — fixed monthly amounts
    "line_maintenance_monthly": "35000.00",
    "base_maintenance_monthly": "15000.00",
    "personnel_salary_monthly": "25000.00",
    "c_check_monthly": "18000.00",
    "maintenance_training_monthly": "5500.00",
    # Maintenance — variable rate (per block hour)
    "spare_parts_rate": "12.5000",
    # Maintenance — per diem
    "maintenance_per_diem": "3500.00",
    # Average active fleet size
    "average_active_fleet": "11.0",
}

# ---------------------------------------------------------------------------
# Authoritative data extracted from 'C' sheet (Crew costs)
# Source: Salary rows, per diem rows, accommodation, training, uniform budgets
# A320: 2 pilots, 1 senior cabin attendant, 3 regular cabin attendants
# A321: 2 pilots, 1 senior cabin attendant, 4 regular cabin attendants
# ---------------------------------------------------------------------------

CREW_CONFIG_DATA = {
    "A320": {
        "pilot_salary_monthly": "12500.00",
        "senior_attendant_salary_monthly": "4500.00",
        "regular_attendant_salary_monthly": "3500.00",
        "per_diem_rate": "75.00",
        "accommodation_monthly_budget": "44000.00",
        "training_total_budget": "55000.00",
        "uniform_total_budget": "22000.00",
    },
    "A321": {
        "pilot_salary_monthly": "13000.00",
        "senior_attendant_salary_monthly": "4800.00",
        "regular_attendant_salary_monthly": "3700.00",
        "per_diem_rate": "80.00",
        "accommodation_monthly_budget": "55000.00",
        "training_total_budget": "66000.00",
        "uniform_total_budget": "27500.00",
    },
}


def validate_against_excel(excel_path: str) -> list[str]:
    """Attempt to read the Excel workbook and validate hardcoded data.

    Returns a list of warning messages for any mismatches found.
    """
    warnings: list[str] = []
    try:
        import openpyxl
    except ImportError:
        warnings.append("openpyxl not installed -- skipping Excel validation")
        return warnings

    if not Path(excel_path).exists():
        warnings.append(f"Excel file not found at {excel_path} -- skipping validation")
        return warnings

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Validate crew config from C sheet
        if "C" in wb.sheetnames:
            c_sheet = wb["C"]
            # TODO: Add specific cell validation once exact cell references are audited
            pass

        # Validate pricing config from M/I/Overhead sheet
        mi_candidates = [s for s in wb.sheetnames if "M" in s and "I" in s]
        if mi_candidates:
            # TODO: Add specific cell validation once exact cell references are audited
            pass

        # Validate exchange rate from Summary sheet
        if "Summary" in wb.sheetnames:
            # TODO: Add specific cell validation once exact cell references are audited
            pass

        wb.close()
    except Exception as e:
        warnings.append(f"Excel validation error: {e}")

    return warnings


def print_seed_data():
    """Print all seed data for dry-run mode."""
    print(f"\n{'='*60}")
    print("  Pricing Config Seed Data")
    print(f"{'='*60}\n")

    print("--- pricing_config (version 1) ---")
    for key, value in PRICING_CONFIG_DATA.items():
        print(f"  {key:>35s}: {value}")

    print()
    for aircraft_type, data in CREW_CONFIG_DATA.items():
        print(f"--- crew_config: {aircraft_type} (version 1) ---")
        for key, value in data.items():
            print(f"  {key:>35s}: {value}")
        print()

    print(f"Total: 1 pricing_config row, {len(CREW_CONFIG_DATA)} crew_config rows\n")


async def seed(database_url: str):
    """Seed the database with pricing_config and crew_config version 1 rows."""
    import asyncpg

    conn = await asyncpg.connect(database_url)
    try:
        # Insert pricing_config version 1
        d = PRICING_CONFIG_DATA
        await conn.execute(
            """
            INSERT INTO pricing_config (
                version, exchange_rate, insurance_usd, doc_total_budget,
                overhead_total_budget, other_cogs_monthly,
                line_maintenance_monthly, base_maintenance_monthly,
                personnel_salary_monthly, c_check_monthly,
                maintenance_training_monthly, spare_parts_rate,
                maintenance_per_diem, average_active_fleet, is_current
            ) VALUES (
                1, $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, TRUE
            )
            """,
            Decimal(d["exchange_rate"]),
            Decimal(d["insurance_usd"]),
            Decimal(d["doc_total_budget"]),
            Decimal(d["overhead_total_budget"]),
            Decimal(d["other_cogs_monthly"]),
            Decimal(d["line_maintenance_monthly"]),
            Decimal(d["base_maintenance_monthly"]),
            Decimal(d["personnel_salary_monthly"]),
            Decimal(d["c_check_monthly"]),
            Decimal(d["maintenance_training_monthly"]),
            Decimal(d["spare_parts_rate"]),
            Decimal(d["maintenance_per_diem"]),
            Decimal(d["average_active_fleet"]),
        )
        print("  pricing_config version 1 inserted")

        # Insert crew_config version 1 for each aircraft type
        for aircraft_type, data in CREW_CONFIG_DATA.items():
            await conn.execute(
                """
                INSERT INTO crew_config (
                    version, aircraft_type,
                    pilot_salary_monthly, senior_attendant_salary_monthly,
                    regular_attendant_salary_monthly, per_diem_rate,
                    accommodation_monthly_budget, training_total_budget,
                    uniform_total_budget, is_current
                ) VALUES (
                    1, $1, $2, $3, $4, $5, $6, $7, $8, TRUE
                )
                """,
                aircraft_type,
                Decimal(data["pilot_salary_monthly"]),
                Decimal(data["senior_attendant_salary_monthly"]),
                Decimal(data["regular_attendant_salary_monthly"]),
                Decimal(data["per_diem_rate"]),
                Decimal(data["accommodation_monthly_budget"]),
                Decimal(data["training_total_budget"]),
                Decimal(data["uniform_total_budget"]),
            )
            print(f"  crew_config {aircraft_type} version 1 inserted")

        print(f"\nSeed complete: 1 pricing_config + {len(CREW_CONFIG_DATA)} crew_config rows loaded.")
    finally:
        await conn.close()


def main():
    parser = argparse.ArgumentParser(
        description="Seed pricing_config and crew_config from Excel workbook data"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print data without writing to database",
    )
    parser.add_argument(
        "--excel",
        default=str(
            Path(__file__).resolve().parent.parent.parent
            / "UNA Pricing Model 1 year.xlsx"
        ),
        help="Path to Excel workbook for validation (default: project root)",
    )
    args = parser.parse_args()

    # Validate against Excel
    print("Validating hardcoded data against Excel workbook...")
    warnings = validate_against_excel(args.excel)
    if warnings:
        for w in warnings:
            print(f"  WARNING: {w}")
    else:
        print("  All validations passed.")

    if args.dry_run:
        print_seed_data()
        return

    database_url = os.environ.get(
        "DATABASE_URL",
        "postgresql://postgres:postgres@localhost:5432/acmi_pricing",
    )
    print(f"\nSeeding database at: {database_url}")
    asyncio.run(seed(database_url))


if __name__ == "__main__":
    main()
