"""Seed script: populate aircraft database from Excel workbook data.

Reads the UNA Pricing Model 1 year.xlsx workbook and populates the aircraft,
aircraft_rates, and epr_matrix_rows tables with all 11 MSNs.

Usage:
    python fastapi-project/scripts/seed_aircraft.py              # Live insert
    python fastapi-project/scripts/seed_aircraft.py --dry-run    # Print only

The script uses hardcoded data extracted from the Excel workbook audit (rows
81-93 for USD rates, rows 14-16 for escalation, and per-MSN EPR tables).
It also attempts to read the Excel file to validate the hardcoded data and
warns if mismatches are found.

All numeric values are wrapped in Decimal(str(value)) to preserve precision.
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from decimal import Decimal
from pathlib import Path

# ---------------------------------------------------------------------------
# Authoritative data extracted from 'A ' sheet of UNA Pricing Model 1 year.xlsx
# Source: Rows 81-93 (USD summary table)
# ---------------------------------------------------------------------------

AIRCRAFT_DATA = {
    3055: {
        "lease_rent": "185000",
        "six_year": "15413.95",
        "twelve_year": "8418.19",
        "ldg": "4333.21",
        "apu": "59.74",
        "llp1": "317.905",
        "llp2": "317.905",
    },
    3378: {
        "lease_rent": "185000",
        "six_year": "14334.75",
        "twelve_year": "7513.065",
        "ldg": "4864.5",
        "apu": "53.82",
        "llp1": "341.775",
        "llp2": "353.13495",
    },
    3461: {
        "lease_rent": "185000",
        "six_year": "16207.16",
        "twelve_year": "8913.47",
        "ldg": "6010.00",
        "apu": "63.02",
        "llp1": "317.35",
        "llp2": "317.35",
    },
    3570: {
        "lease_rent": "200000",
        "six_year": "14729.085",
        "twelve_year": "8248.95",
        "ldg": "5302.305",
        "apu": "52.785",
        "llp1": "341.775",
        "llp2": "341.775",
    },
    3605: {
        "lease_rent": "185000",
        "six_year": "14965.48",
        "twelve_year": "8173.60",
        "ldg": "4207.00",
        "apu": "58.35",
        "llp1": "317.905",
        "llp2": "317.905",
    },
    4247: {
        "lease_rent": "185000",
        "six_year": "15414.465",
        "twelve_year": "8419.323",
        "ldg": "4333.725",
        "apu": "60.049",
        "llp1": "360.22",
        "llp2": "360.22",
    },
    5228: {
        "lease_rent": "210000",
        "six_year": "16207.16",
        "twelve_year": "8913.47",
        "ldg": "6010.00",
        "apu": "63.02",
        "llp1": "353.29",
        "llp2": "353.29",
    },
    5931: {
        "lease_rent": "220000",
        "six_year": "16207.16",
        "twelve_year": "8913.47",
        "ldg": "6010.00",
        "apu": "63.02",
        "llp1": "317.35",
        "llp2": "317.35",
    },
    1932: {
        "lease_rent": "235000",
        "six_year": "17139.6",
        "twelve_year": "10712.25",
        "ldg": "5892.255",
        "apu": "69.345",
        "llp1": "317.905",
        "llp2": "317.905",
    },
    1960: {
        "lease_rent": "235000",
        "six_year": "17139.6",
        "twelve_year": "10712.25",
        "ldg": "5892.255",
        "apu": "69.345",
        "llp1": "317.905",
        "llp2": "317.905",
    },
    1503: {
        "lease_rent": "235000",
        "six_year": "17139.6",
        "twelve_year": "10712.25",
        "ldg": "5892.255",
        "apu": "69.345",
        "llp1": "317.905",
        "llp2": "317.905",
    },
}

# ---------------------------------------------------------------------------
# Escalation rates from rows 14-16
# Stored as decimal fractions (0.05 = 5%)
# ---------------------------------------------------------------------------

ESCALATION_RATES = {
    3055: {"epr": "0.05", "llp": "0.08", "af_apu": "0.03"},
    3378: {"epr": "0.03", "llp": "0.085", "af_apu": "0.035"},
    3461: {"epr": "0.05", "llp": "0.08", "af_apu": "0.03"},
    3570: {"epr": "0.05", "llp": "0.08", "af_apu": "0.03"},
    3605: {"epr": "0.05", "llp": "0.08", "af_apu": "0.03"},
    4247: {"epr": "0.045", "llp": "0.08", "af_apu": "0.03"},
    5228: {"epr": "0.035", "llp": "0.08", "af_apu": "0.035"},
    5931: {"epr": "0", "llp": "0", "af_apu": "0"},
    1932: {"epr": "0.05", "llp": "0.08", "af_apu": "0.04"},
    1960: {"epr": "0.05", "llp": "0.08", "af_apu": "0.04"},
    1503: {"epr": "0.05", "llp": "0.08", "af_apu": "0.04"},
}

# ---------------------------------------------------------------------------
# EPR tables per MSN — (cycle_ratio, benign_rate, hot_rate)
# Using 2026 tables where both 2025/2026 exist (MSNs 1932, 1960, 1503)
# ---------------------------------------------------------------------------

EPR_TABLES = {
    3055: [
        # Rows 36-41 cols B,D,E
        ("1.0", "448.22", "672.33"),
        ("1.5", "319.07", "478.61"),
        ("2.0", "273.49", "410.23"),
        ("2.25", "258.30", "387.44"),
        ("2.5", "250.70", "376.05"),
        ("3.0", "243.10", "364.65"),
    ],
    3378: [
        # Rows 36-50 cols G,I,J
        ("0.62", "669.32", "870.73"),
        ("0.80", "576.65", "750.79"),
        ("0.87", "540.61", "704.15"),
        ("1.12", "452.78", "589.07"),
        ("1.37", "392.20", "510.32"),
        ("1.62", "354.35", "461.86"),
        ("1.87", "325.57", "424.01"),
        ("2.12", "302.86", "393.71"),
        ("2.37", "284.69", "371.01"),
        ("2.62", "272.58", "354.35"),
        ("2.87", "265.00", "345.26"),
        ("3.12", "261.97", "340.72"),
        ("3.37", "260.46", "339.21"),
        ("3.62", "258.94", "337.69"),
        ("4.0", "255.92", "333.15"),
    ],
    3461: [
        # Rows 37-47 cols L,N,O
        ("0.99", "507.15", "659.40"),
        ("1.24", "426.30", "553.35"),
        ("1.49", "365.40", "474.60"),
        ("1.74", "304.50", "395.85"),
        ("1.99", "292.95", "381.15"),
        ("2.24", "282.45", "367.50"),
        ("2.49", "270.90", "351.75"),
        ("2.74", "263.55", "342.30"),
        ("2.99", "257.25", "333.90"),
        ("3.24", "249.90", "325.50"),
        ("3.5", "243.60", "316.05"),
    ],
    3570: [
        # Rows 36-44 cols Q,S,T
        ("0.5", "645.09", "838.62"),
        ("0.75", "588.94", "765.63"),
        ("1.0", "532.80", "692.64"),
        ("1.5", "420.51", "546.66"),
        ("2.0", "357.57", "464.84"),
        ("2.25", "328.15", "426.60"),
        ("2.5", "311.96", "405.55"),
        ("3.0", "283.90", "369.06"),
        ("3.5", "268.11", "348.54"),
    ],
    3605: [
        # Rows 70-75 cols Q,S,T
        ("1.0", "448.58", "672.87"),
        ("1.5", "319.79", "479.69"),
        ("2.0", "273.49", "410.23"),
        ("2.25", "259.02", "388.53"),
        ("2.5", "251.78", "377.68"),
        ("3.0", "243.10", "364.65"),
    ],
    4247: [
        # Rows 70-74 cols G,I,J
        ("1.0", "417.38", "696.13"),
        ("1.5", "310.05", "515.76"),
        ("2.0", "263.84", "439.74"),
        ("2.25", "238.50", "396.51"),
        ("2.5", "228.07", "380.12"),
    ],
    5228: [
        # Rows 70-80 cols L,N,O
        ("0.99", "450.00", "630.00"),
        ("1.24", "400.00", "560.00"),
        ("1.49", "335.00", "469.00"),
        ("1.74", "315.00", "441.00"),
        ("1.99", "295.00", "413.00"),
        ("2.24", "275.00", "385.00"),
        ("2.49", "265.00", "371.00"),
        ("2.74", "250.00", "350.00"),
        ("2.99", "235.00", "329.00"),
        ("3.24", "220.00", "308.00"),
        ("3.5", "210.00", "294.00"),
    ],
    5931: [
        # Rows 88-98 cols V,W,X
        ("0.99", "507.15", "710.01"),
        ("1.24", "426.30", "596.82"),
        ("1.49", "365.40", "511.56"),
        ("1.74", "304.50", "426.30"),
        ("1.99", "285.00", "399.00"),
        ("2.24", "275.00", "385.00"),
        ("2.49", "262.00", "366.80"),
        ("2.74", "250.00", "350.00"),
        ("2.99", "235.00", "329.00"),
        ("3.24", "220.00", "308.00"),
        ("3.5", "200.00", "280.00"),
    ],
    1932: [
        # 2026 table: Rows 38-49 cols V,W,X
        ("0.86", "656.21", "919.50"),
        ("1.11", "528.56", "740.62"),
        ("1.36", "451.54", "632.71"),
        ("1.61", "398.79", "557.32"),
        ("1.86", "358.70", "502.62"),
        ("2.11", "328.11", "458.27"),
        ("2.36", "302.79", "424.27"),
        ("2.61", "281.69", "396.18"),
        ("2.86", "264.81", "371.05"),
        ("3.11", "251.09", "351.83"),
        ("3.5", "237.38", "332.62"),
        ("4.0", "237.38", "332.62"),
    ],
    1960: [
        # 2026 table: Rows 38-50 cols Z,AA,AB
        ("0.86", "656.21", "853.07"),
        ("1.11", "528.56", "687.12"),
        ("1.36", "451.54", "587.00"),
        ("1.61", "398.79", "518.43"),
        ("1.86", "358.70", "466.31"),
        ("2.11", "328.11", "426.54"),
        ("2.36", "302.79", "393.62"),
        ("2.61", "281.69", "366.19"),
        ("2.86", "264.81", "344.25"),
        ("3.11", "251.09", "326.42"),
        ("3.5", "237.38", "308.59"),
        ("3.99", "220.14", "286.19"),
        ("4.0", "219.79", "285.73"),
    ],
    1503: [
        # 2026 table: Rows 72-82 cols V,W,X
        ("0.86", "704.33", "950.95"),
        ("1.11", "568.48", "767.03"),
        ("1.36", "485.93", "656.26"),
        ("1.61", "428.45", "577.89"),
        ("1.86", "385.61", "520.41"),
        ("2.11", "353.21", "476.52"),
        ("2.36", "326.04", "438.90"),
        ("2.61", "303.05", "408.60"),
        ("2.86", "284.24", "382.47"),
        ("3.11", "269.61", "364.71"),
        ("3.5", "256.03", "344.85"),
    ],
}

# All MSNs in fleet order
ALL_MSNS = [3055, 3378, 3461, 3570, 3605, 4247, 5228, 5931, 1932, 1960, 1503]


def validate_against_excel(excel_path: str) -> list[str]:
    """Attempt to read the Excel workbook and validate hardcoded data.

    Returns a list of warning messages for any mismatches found.
    """
    warnings = []
    try:
        import openpyxl
    except ImportError:
        warnings.append("openpyxl not installed -- skipping Excel validation")
        return warnings

    if not Path(excel_path).exists():
        warnings.append(f"Excel file not found at {excel_path} -- skipping validation")
        return warnings

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb["A "]

        # Validate MSN list from row 6, cols C-M
        excel_msns = []
        for col in range(3, 14):
            val = sheet.cell(row=6, column=col).value
            if val is not None:
                excel_msns.append(int(val))

        if set(excel_msns) != set(ALL_MSNS):
            warnings.append(
                f"MSN mismatch: Excel has {excel_msns}, hardcoded has {ALL_MSNS}"
            )

        # Validate escalation rates from rows 14-16
        esc_labels = {14: "epr", 15: "llp", 16: "af_apu"}
        for col_idx, msn in enumerate(excel_msns):
            if msn not in ESCALATION_RATES:
                continue
            for row, esc_key in esc_labels.items():
                excel_val = sheet.cell(row=row, column=col_idx + 3).value
                hardcoded_val = float(ESCALATION_RATES[msn][esc_key])
                if excel_val is not None and abs(float(excel_val) - hardcoded_val) > 0.001:
                    warnings.append(
                        f"Escalation mismatch MSN {msn} {esc_key}: "
                        f"Excel={excel_val}, hardcoded={hardcoded_val}"
                    )

        wb.close()
    except Exception as e:
        warnings.append(f"Excel validation error: {e}")

    return warnings


def print_seed_data():
    """Print all seed data for dry-run mode."""
    print(f"\n{'='*60}")
    print(f"  Aircraft Seed Data — {len(ALL_MSNS)} MSNs")
    print(f"{'='*60}\n")

    for msn in ALL_MSNS:
        data = AIRCRAFT_DATA[msn]
        esc = ESCALATION_RATES[msn]
        epr = EPR_TABLES.get(msn, [])

        print(f"--- MSN {msn} (A320) ---")
        print(f"  Lease Rent:   ${data['lease_rent']}")
        print(f"  6Y Check:     ${data['six_year']}")
        print(f"  12Y Check:    ${data['twelve_year']}")
        print(f"  LDG:          ${data['ldg']}")
        print(f"  APU Rate:     ${data['apu']}")
        print(f"  LLP1 Rate:    ${data['llp1']}")
        print(f"  LLP2 Rate:    ${data['llp2']}")
        print(f"  Escalation:   EPR={esc['epr']}, LLP={esc['llp']}, AF+APU={esc['af_apu']}")
        print(f"  EPR Matrix:   {len(epr)} rows")
        for ratio, benign, hot in epr:
            print(f"    CR={ratio:>5s}  Benign={benign:>8s}  Hot={hot:>8s}")
        print()

    print(f"Total: {len(ALL_MSNS)} aircraft, "
          f"{sum(len(EPR_TABLES.get(m, [])) for m in ALL_MSNS)} EPR rows\n")


async def seed(database_url: str):
    """Seed the database with all 11 aircraft and their cost data."""
    import asyncpg

    conn = await asyncpg.connect(database_url)
    try:
        for msn in ALL_MSNS:
            data = AIRCRAFT_DATA[msn]
            esc = ESCALATION_RATES[msn]
            epr_rows = EPR_TABLES.get(msn, [])

            # Upsert aircraft
            row = await conn.fetchrow(
                "INSERT INTO aircraft (msn, aircraft_type) VALUES ($1, $2) "
                "ON CONFLICT (msn) DO UPDATE SET updated_at = NOW() "
                "RETURNING id",
                msn,
                "A320",
            )
            aircraft_id = row["id"]
            print(f"  Aircraft MSN {msn} -> id={aircraft_id}")

            # Upsert rates
            await conn.execute(
                """
                INSERT INTO aircraft_rates (
                    aircraft_id, lease_rent_usd, six_year_check_usd,
                    twelve_year_check_usd, ldg_usd, apu_rate_usd,
                    llp1_rate_usd, llp2_rate_usd,
                    epr_escalation, llp_escalation, af_apu_escalation
                ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)
                ON CONFLICT (aircraft_id) DO UPDATE SET
                    lease_rent_usd = EXCLUDED.lease_rent_usd,
                    six_year_check_usd = EXCLUDED.six_year_check_usd,
                    twelve_year_check_usd = EXCLUDED.twelve_year_check_usd,
                    ldg_usd = EXCLUDED.ldg_usd,
                    apu_rate_usd = EXCLUDED.apu_rate_usd,
                    llp1_rate_usd = EXCLUDED.llp1_rate_usd,
                    llp2_rate_usd = EXCLUDED.llp2_rate_usd,
                    epr_escalation = EXCLUDED.epr_escalation,
                    llp_escalation = EXCLUDED.llp_escalation,
                    af_apu_escalation = EXCLUDED.af_apu_escalation,
                    updated_at = NOW()
                """,
                aircraft_id,
                Decimal(data["lease_rent"]),
                Decimal(data["six_year"]),
                Decimal(data["twelve_year"]),
                Decimal(data["ldg"]),
                Decimal(data["apu"]),
                Decimal(data["llp1"]),
                Decimal(data["llp2"]),
                Decimal(esc["epr"]),
                Decimal(esc["llp"]),
                Decimal(esc["af_apu"]),
            )

            # Upsert EPR matrix rows
            for ratio, benign, hot in epr_rows:
                await conn.execute(
                    """
                    INSERT INTO epr_matrix_rows (aircraft_id, cycle_ratio, benign_rate, hot_rate)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (aircraft_id, cycle_ratio) DO UPDATE SET
                        benign_rate = EXCLUDED.benign_rate,
                        hot_rate = EXCLUDED.hot_rate
                    """,
                    aircraft_id,
                    Decimal(ratio),
                    Decimal(benign),
                    Decimal(hot),
                )

            print(f"    Rates and {len(epr_rows)} EPR rows inserted")

        print(f"\nSeed complete: {len(ALL_MSNS)} aircraft loaded.")
    finally:
        await conn.close()


def main():
    parser = argparse.ArgumentParser(description="Seed aircraft database from Excel workbook data")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print data without writing to database",
    )
    parser.add_argument(
        "--excel",
        default=str(Path(__file__).resolve().parent.parent.parent / "UNA Pricing Model 1 year.xlsx"),
        help="Path to Excel workbook for validation (default: project root)",
    )
    args = parser.parse_args()

    # Validate against Excel
    print("Validating hardcoded data against Excel workbook...")
    warnings = validate_against_excel(args.excel)
    if warnings:
        for w in warnings:
            print(f"  WARNING: {w}")
    else:
        print("  All validations passed.")

    if args.dry_run:
        print_seed_data()
        return

    database_url = os.environ.get(
        "DATABASE_URL",
        "postgresql://postgres:postgres@localhost:5432/acmi_pricing",
    )
    print(f"\nSeeding database at: {database_url}")
    asyncio.run(seed(database_url))


if __name__ == "__main__":
    main()
